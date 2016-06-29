#!/usr/bin/env python3

import numpy as np
import openpyxl
import os, sys
from log import frlog


class IO_Array:
    """Módulo que importa/exporta os dados dos espectros de massas.
    """

    def __init__(self, myfile):
        # Arquivo que será utilizado
        self.myfile = myfile
    
    def import_file(self, method):
        """Função que abre o arquivo excel que será utilizado para importar/exportar os dados.
        """
        if method == 'create':
            # Verifica se o arquivo existe
            if os.path.isfile(self.myfile):
                # Abre o arquivo excel existente. 
                self.wb = openpyxl.load_workbook(self.myfile)
                self.sheet = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
            else:
                print('Arquivo não encontrado!')
                #sys.exit()
        elif method == 'export':
            # Cria o objeto de um arquivo excel para o salvamento.
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
            self.sheet.title = 'Sheet1'
        else:
            print('Erro interno - import_file')
            #sys.exit()
            
    def create_array(self, controller):
        """Função que retorna os dados do arquivo aberto, excel ou texto.
        """

        self.controller = controller
        # Verifica o tipo de arquivo
        if self.myfile.split('.')[-1] == 'xlsx':
            # Se for xlsx chama a função import_file que retorna o objeto excel do arquivo.
            self.import_file('create')
            # Verifica o tamanho do arquivo.
            self.ncol = self.sheet.max_column
            self.nrow = self.sheet.max_row
            # Cria uma lista vazia.
            self.mylist=[]
            # Itera nas linhas
            for i in range(self.nrow):
                # Para cada linha cria uma nova lista com as colunas
                self.list=[]
                # Itera nas colunas
                for j in range(self.ncol):
                    # Acrescenta os valores de cada coluna em self.list
                    self.list.append(self.sheet.cell(row=i+1, column=j+1).value)
                # Acrescenta cada self.list como uma nova linha
                self.mylist.append(self.list)        
        else:
            # Caso o arquivo não seja xlsx tenta tratá-lo como texto
            try:
                # Cria lista vazia
                self.mylist=[]
                # Abre arquivo em modo texto
                f = open(self.myfile)
                # Adiciona cada linha na lista
                for line in f:
                    self.mylist.append(line)
            # Se não conseguir ler o arquivo como texto retorna o erro de arquivo não reconhecido
            except:
                self.controller.frames[frlog.FrameLog].WriteLog('error', 'Tipo de arquivo não reconhecido!')
                # Esvazia a lista
                self.mylist = None
                #sys.exit()
        # Retorna a lista ao programa principal
        return self.mylist

    def export_array(self, mylist):
        """Função que despeja o conteúdo da lista num arquivo excel e o salva.
        """

        # Transforma a lista num array numpy - facilita na transferência
        self.mylist = np.array(mylist)
        # Chama a função import que retorna o objeto do arquivo excel 
        self.import_file('export')
        # Verifica quantas dimensões tem o array (1 ou 2)
        if len(self.mylist.shape) > 1:
            # Verifica as dimensões do array
            self.nrow, self.ncol = self.mylist.shape
            # e itera sobre elas
            for i in range(self.nrow):
                for j in range(self.ncol):
                    # passando os valores da lista como floats
                    self.sheet.cell(row=i+1, column=j+1).value = float(self.mylist[i, j])
        else:
            # Faz o mesmo para um array unidimensional
            self.nrow = len(self.mylist)
            for i in range(self.nrow):
                self.sheet.cell(row=i+1, column=1).value = float(self.mylist[i])
        # Salva o arquivo excel
        self.wb.save(self.myfile)

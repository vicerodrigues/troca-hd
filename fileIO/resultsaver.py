import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Series, Reference
import os, sys
from aux import combin


class OutputResults:
	"""Módulo que salva os dados de saída (resultados) da simulação em um arquivo excel .xlsx
	"""

	def __init__(self, myfile):
		# Recebe o nome do arquivo a ser salvo
		self.myfile = myfile

	def export_results(self, mylist, args):
		"""Função que estrutura e salva o arquivo excel com os resultados.
		"""

		# Transforma a lista de resultados num array numpy para facilitar a escrita no arquivo
		self.mylist = np.array(mylist)
		# args refere-se aos parâmetros que descrevem o resultado da simulação - R² etc
		self.args = args

		# Retorna o conjunto de objetos do arquivo excel
		self.wb = openpyxl.Workbook()
		self.sheet = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
		self.sheet.title = 'Sheet1'

		# Salva os resultados e labels em suas respectivas células.
		self.sheet.cell(row=1, column=1).value = 'Hidrocarboneto'
		self.sheet.cell(row=1, column=2).value = 'Experimental'
		self.sheet.cell(row=1, column=3).value = 'Binomial'
		self.sheet.cell(row=1, column=5).value = 'Phi'
		self.sheet.cell(row=1, column=6).value = self.args[0]
		self.sheet.cell(row=2, column=5).value = 'MDC (%)'
		self.sheet.cell(row=2, column=6).value = self.args[1]
		self.sheet.cell(row=3, column=5).value = 'R^2'
		self.sheet.cell(row=3, column=6).value = self.args[2]
		self.sheet.cell(row=4, column=5).value = 'Método'
		if self.args[3] == 1:
			self.sheet.cell(row=4, column=6).value = 'Least-squares'
		elif self.args[3] == 2:
			self.sheet.cell(row=4, column=6).value = 'Non-negative least-squares'

		# Calcula a distribuição binomial com base no MDC fracionário
		self.nrow = len(self.mylist)
		for i in range(self.nrow):
			self.sheet.cell(row=i+2, column=1).value = 'D%i' %i
			self.sheet.cell(row=i+2, column=2).value = float(self.mylist[i])
			self.binom = 100 * combin.Combinations(len(self.mylist)-1, i).calcComb * (self.args[1]/100)**i *\
						 (1-self.args[1]/100)**((len(self.mylist)-1)-i)
			self.sheet.cell(row=i+2, column=3).value = self.binom

		# Cria o gráfico de comparação no arquivo excel
		self.chart = BarChart()
		self.chart.type = "col"
		self.chart.style = 10
		self.chart.title = "Comparação Experimental x Binomial"
		self.chart.y_axis.title = 'Teor (%)'
		#self.chart.x_axis.title = 'Di' - Não utilizado pois estava rotacionado de 90º

		# Dados
		self.data = Reference(self.sheet, min_col=2, min_row=1, max_row=8, max_col=3)
		# Categorias
		self.cats = Reference(self.sheet, min_col=1, min_row=2, max_row=8)
		# Adicionando dados com títulos
		self.chart.add_data(self.data, titles_from_data=True)
		self.chart.set_categories(self.cats)
		self.chart.shape = 4
		self.sheet.add_chart(self.chart, "E7")
		# Salvando o arquivo excel de fato
		self.wb.save(self.myfile)



#!/usr/bin/env python3

import numpy as np
import openpyxl
import os, sys

class IO_Array:
    def __init__(self, myfile):
        self.myfile = myfile
    
    def import_file(self, method):
        if method == 'create':
            if os.path.isfile(self.myfile):
                self.wb = openpyxl.load_workbook(self.myfile)
                self.sheet = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
            else:
                print('Arquivo não encontrado!')
                sys.exit()
        elif method == 'export':
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
            self.sheet.title = 'Sheet1'
        else:
            print('Erro interno - import_file')
            sys.exit()
            
    def create_array(self):
        if self.myfile.split('.')[-1] == 'xlsx':
            self.import_file('create')
            self.ncol = self.sheet.max_column
            self.nrow = self.sheet.max_row
            self.mylist=[]
            for i in range(self.nrow):
                self.list=[]
                for j in range(self.ncol):
                    self.list.append(self.sheet.cell(row=i+1, column=j+1).value)
                self.mylist.append(self.list)        
            #self.myarray = np.array(self.mylist)
        else:
            try:
                self.mylist=[]
                f = open(self.myfile)
                for line in f:
                    self.mylist.append(line)
            except:
                print('Tipo de arquivo não reconhecido!')
                sys.exit()
        return self.mylist

    def export_array(self, mylist):
        self.mylist = np.array(mylist)
        self.import_file('export')
        if len(self.mylist.shape) > 1:
            self.nrow, self.ncol = self.mylist.shape
            for i in range(self.nrow):
                for j in range(self.ncol):
                    self.sheet.cell(row=i+1, column=j+1).value = float(self.mylist[i, j])
        else:
            self.nrow = len(self.mylist)
            for i in range(self.nrow):
                self.sheet.cell(row=i+1, column=1).value = float(self.mylist[i])

        self.wb.save(self.myfile)
